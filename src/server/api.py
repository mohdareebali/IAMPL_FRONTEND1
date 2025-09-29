# -*- coding: utf-8 -*-
"""
Merged Flask backend:
- MySQL auth (company/employee) + email/OTP flows
- OCR/Extraction (PDF/Images/Docx/Excel)
- CMM Excel parser
- Supplier autocomplete

Run:  python app.py
"""

import os
import io
import re
import csv
import json
import random
from typing import List
from datetime import datetime, timedelta

from flask import Flask, request, jsonify, Response, stream_with_context
from flask_cors import CORS

# --- OCR / File processing deps
import pytesseract
from PIL import Image, ImageEnhance
import fitz  # PyMuPDF
import docx
import openpyxl

# --- Email + DB
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import mysql.connector

# ==============================
# 🚀 FLASK APP INIT
# ==============================
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # allow up to 20 MB uploads
CORS(app)

# --- OPTIONAL: Serve a built frontend (set FRONTEND_BUILD_PATH env var to enable)
FRONTEND_BUILD_PATH = os.environ.get("FRONTEND_BUILD_PATH", "")
if FRONTEND_BUILD_PATH and os.path.isdir(FRONTEND_BUILD_PATH):
    app.static_folder = FRONTEND_BUILD_PATH

    @app.route('/', defaults={'path': ''})
    @app.route('/<path:path>')
    def serve_frontend(path):
        # serve exact file if it exists, else fallback to index.html (SPA)
        if path and os.path.exists(os.path.join(app.static_folder, path)):
            return app.send_static_file(path)
        return app.send_static_file('index.html')

# ==============================
# ✅ MySQL Connection
# ==============================
db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="areeb@123",
    database="innovascape"
)
cursor = db.cursor(dictionary=True)

# ==============================
# 🚀 Email Transporter (Gmail app password)
# ==============================
EMAIL_USER = "aribali1804@gmail.com"
EMAIL_PASS = "xpfp szcf fcjl grsw"  # Gmail app password

def send_email(to, subject, text, html=None):
    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = EMAIL_USER
        msg["To"] = to
        msg["Subject"] = subject

        part1 = MIMEText(text, "plain")
        msg.attach(part1)
        if html:
            part2 = MIMEText(html, "html")
            msg.attach(part2)

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASS)
            server.sendmail(EMAIL_USER, to, msg.as_string())

        return True
    except Exception as e:
        print("❌ Email error:", e)
        return False


# ==============================
# 🔐 COMPANY ROUTES
# ==============================
@app.route("/api/register", methods=["POST"])
def register_company():
    data = request.json
    companyName = data.get("companyName")
    email = data.get("email")
    companyId = data.get("companyId")
    password = data.get("password")

    if not all([companyName, email, companyId, password]):
        return jsonify({"error": "All fields are required"}), 400

    try:
        cursor.execute(
            "INSERT INTO companies (company_name, email, company_id, password) VALUES (%s, %s, %s, %s)",
            (companyName, email, companyId, password),
        )
        db.commit()
        return jsonify({"message": "✅ Company registered successfully!", "companyId": companyId})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/login", methods=["POST"])
def company_login():
    data = request.json
    cid = data.get("id")
    password = data.get("password")

    if not cid or not password:
        return jsonify({"error": "ID and Password required"}), 400

    cursor.execute("SELECT * FROM companies WHERE company_id=%s OR email=%s", (cid, cid))
    company = cursor.fetchone()

    if not company or password != company["password"]:
        return jsonify({"error": "Invalid credentials"}), 401

    return jsonify({"message": "✅ Login successful!", "company": company})


# ==============================
# 👤 EMPLOYEE ROUTES
# ==============================
@app.route("/api/employees", methods=["POST"])
def add_employee():
    data = request.json
    employee_id = data.get("employee_id")
    email = data.get("email")
    password = data.get("password")
    company_id = data.get("company_id")

    if not all([employee_id, email, password, company_id]):
        return jsonify({"error": "All fields required"}), 400

    cursor.execute("SELECT * FROM companies WHERE company_id=%s", (company_id,))
    if not cursor.fetchone():
        return jsonify({"error": "❌ Invalid Company ID"}), 400

    try:
        cursor.execute(
            "INSERT INTO employees (employee_id, email, password, company_id, created_at) VALUES (%s, %s, %s, %s, NOW())",
            (employee_id, email, password, company_id),
        )
        db.commit()
        return jsonify({"message": "✅ Employee added successfully!", "employeeId": employee_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/employees", methods=["GET"])
def get_employees():
    cursor.execute("SELECT employee_id, email, company_id, password, created_at, password_changed_at FROM employees")
    employees = cursor.fetchall()

    for emp in employees:
        if emp.get("created_at"):
            emp["created_at"] = emp["created_at"].strftime("%Y-%m-%d %H:%M:%S")
        else:
            emp["created_at"] = "N/A"

        if emp.get("password_changed_at"):
            emp["password_changed_at"] = emp["password_changed_at"].strftime("%Y-%m-%d %H:%M:%S")
        else:
            emp["password_changed_at"] = "N/A"

    return jsonify(employees)


@app.route("/api/employees/check/<string:employee_id>", methods=["GET"])
def check_employee_id(employee_id):
    cursor.execute("SELECT employee_id FROM employees WHERE employee_id=%s", (employee_id,))
    exists = cursor.fetchone()
    return jsonify({"exists": bool(exists)})


@app.route("/api/employees/login", methods=["POST"])
def employee_login():
    data = request.json
    identifier = data.get("identifier")
    password = data.get("password")

    if not identifier or not password:
        return jsonify({"error": "Employee ID or Email and Password required"}), 400

    cursor.execute("SELECT * FROM employees WHERE employee_id=%s OR email=%s", (identifier, identifier))
    employee = cursor.fetchone()

    if not employee or password != employee["password"]:
        return jsonify({"error": "Invalid credentials"}), 401

    # strip sensitive fields
    employee.pop("password", None)
    employee.pop("otp", None)
    employee.pop("otp_expiry", None)

    if employee.get("created_at"):
        employee["created_at"] = employee["created_at"].strftime("%Y-%m-%d %H:%M:%S")
    else:
        employee["created_at"] = "N/A"
    if employee.get("password_changed_at"):
        employee["password_changed_at"] = employee["password_changed_at"].strftime("%Y-%m-%d %H:%M:%S")
    else:
        employee["password_changed_at"] = "N/A"

    return jsonify({"message": "✅ Employee login successful!", "employee": employee})


# ==============================
# 🔁 PASSWORD RESET (Employee)
# ==============================
@app.route("/api/employees/forgot-password", methods=["POST"])
def employee_forgot_password():
    data = request.json
    identifier = data.get("identifier")

    cursor.execute("SELECT email, employee_id FROM employees WHERE employee_id=%s OR email=%s", (identifier, identifier))
    employee = cursor.fetchone()

    if not employee:
        return jsonify({"error": "Employee not found"}), 404

    # changed to 5-digit OTP
    otp = random.randint(10000, 99999)
    otp_expiry = datetime.now() + timedelta(minutes=1)

    cursor.execute(
        "UPDATE employees SET otp=%s, otp_expiry=%s WHERE employee_id=%s",
        (otp, otp_expiry, employee["employee_id"])
    )
    db.commit()

    send_email(
        employee["email"],
        "OTP for Password Reset",
        f"Your OTP for password reset is {otp}. It is valid for 1 minute.",
    )

    return jsonify({"message": f"✅ OTP sent to Employee: {employee['employee_id']}"})

# Verify OTP endpoint
@app.route("/api/employees/verify-otp", methods=["POST"])
def employee_verify_otp():
    """
    Verify the 5-digit OTP for an employee (identifier = employee_id or email).
    Returns success if OTP matches and is not expired.
    """
    data = request.json or {}
    identifier = data.get("identifier")
    otp = data.get("otp")

    if not identifier or otp is None:
        return jsonify({"error": "Identifier and OTP are required"}), 400

    cursor.execute("SELECT * FROM employees WHERE employee_id=%s OR email=%s", (identifier, identifier))
    employee = cursor.fetchone()

    if not employee:
        return jsonify({"error": "Employee not found"}), 404

    # Debugging helper (server logs)
    print("VERIFY REQUEST:", {"identifier": identifier, "otp": otp, "db_otp": employee.get("otp"), "db_expiry": employee.get("otp_expiry")})

    # compare as strings to avoid numeric formatting issues (leading zeros, etc.)
    if str(employee.get("otp")) != str(otp).strip():
        return jsonify({"error": "Invalid OTP"}), 400

    if not employee.get("otp_expiry") or datetime.now() > employee["otp_expiry"]:
        return jsonify({"error": "OTP expired"}), 400

    # Optionally clear OTP immediately so user can proceed to reset password
    cursor.execute("UPDATE employees SET otp=NULL, otp_expiry=NULL WHERE employee_id=%s", (employee["employee_id"],))
    db.commit()

    return jsonify({"message": "✅ OTP verified", "employee_id": employee["employee_id"], "email": employee["email"]})

@app.route("/api/employees/reset-password", methods=["POST"])
def employee_reset_password():
    """
    Resets employee password.
    Behaviour:
      - If employee.otp IS NULL in DB (OTP already verified / consumed), allow reset without providing otp.
      - If employee.otp IS NOT NULL, require otp parameter to match and be unexpired.
    """
    data = request.json or {}
    identifier = data.get("identifier")
    otp = data.get("otp")
    newPassword = data.get("newPassword")

    if not identifier or not newPassword:
        return jsonify({"error": "Identifier and newPassword are required"}), 400

    cursor.execute("SELECT * FROM employees WHERE employee_id=%s OR email=%s", (identifier, identifier))
    employee = cursor.fetchone()

    if not employee:
        return jsonify({"error": "Employee not found"}), 404

    db_otp = employee.get("otp")
    db_expiry = employee.get("otp_expiry")

    # If OTP present in DB -> require otp param and validate it
    if db_otp is not None:
        if otp is None:
            return jsonify({"error": "OTP is required"}), 400
        if str(db_otp) != str(otp).strip():
            return jsonify({"error": "Invalid OTP"}), 400
        if not db_expiry or datetime.now() > db_expiry:
            return jsonify({"error": "OTP expired"}), 400

    # If db_otp is None => OTP was already verified (or not set) -> allow reset

    try:
        password_changed_at = datetime.now()
        cursor.execute(
            "UPDATE employees SET password=%s, otp=NULL, otp_expiry=NULL, password_changed_at=%s WHERE employee_id=%s",
            (newPassword, password_changed_at, employee["employee_id"])
        )
        db.commit()
        return jsonify({"message": "✅ Password updated successfully!"})
    except Exception as e:
        print(f"ERROR updating password for {employee['employee_id']}: {e}")
        return jsonify({"error": "Failed to update password"}), 500


# ==============================
# 📤 EMPLOYEE CREDENTIALS SHARE
# ==============================
@app.route("/api/employees/share", methods=["POST"])
def share_employee_credentials():
    data = request.json
    employee_id = data.get("employee_id")
    email = data.get("email")
    password = data.get("password")

    if not all([employee_id, email, password]):
        return jsonify({"error": "Employee ID, Email and Password required"}), 400

    html = f"""
    <div>
        <h2>Welcome to Innovascape 🎉</h2>
        <p>Here are your login credentials:</p>
        <ul>
          <li><b>Employee ID:</b> {employee_id}</li>
          <li><b>Email:</b> {email}</li>
          <li><b>Password:</b> {password}</li>
        </ul>
        <p>⚠️ Keep these credentials safe.</p>
    </div>
    """

    send_email(email, "Your Employee Credentials",
               f"Employee ID: {employee_id}\nEmail: {email}\nPassword: {password}", html)
    return jsonify({"message": f"✅ Credentials sent to {email}"})


# ==============================
# 🔁 PASSWORD RESET (Manager/Company)
# ==============================
@app.route("/api/manager/forgot-password", methods=["POST"])
def manager_forgot_password():
    data = request.json
    identifier = data.get("identifier")

    cursor.execute("SELECT company_id, email FROM companies WHERE company_id=%s OR email=%s", (identifier, identifier))
    company = cursor.fetchone()

    if not company:
        return jsonify({"error": "Company not found"}), 404

    # changed to 5-digit OTP
    otp = random.randint(10000, 99999)
    otp_expiry = datetime.now() + timedelta(minutes=1)

    cursor.execute(
        "UPDATE companies SET otp=%s, otp_expiry=%s WHERE company_id=%s",
        (otp, otp_expiry, company["company_id"])
    )
    db.commit()

    send_email(company["email"], "OTP for Password Reset", f"Your OTP is {otp}. Valid for 1 minute.")

    # Return structured data for frontend convenience
    return jsonify({
        "message": f"✅ OTP sent to Company ID: {company['company_id']}",
        "company_id": company["company_id"],
        "email": company["email"]
    })


@app.route("/api/manager/verify-otp", methods=["POST"])
def manager_verify_otp():
    """Verify 5-digit OTP for a company/manager (identifier = company_id or email)."""
    data = request.json or {}
    identifier = data.get("identifier")
    otp = data.get("otp")

    if not identifier or otp is None:
        return jsonify({"error": "Identifier and OTP are required"}), 400

    cursor.execute("SELECT * FROM companies WHERE company_id=%s OR email=%s", (identifier, identifier))
    company = cursor.fetchone()

    if not company:
        return jsonify({"error": "Company not found"}), 404

    # Debug/logging
    print("MANAGER VERIFY REQUEST:", {"identifier": identifier, "otp": otp, "db_otp": company.get("otp"), "db_expiry": company.get("otp_expiry")})

    if str(company.get("otp")) != str(otp).strip():
        return jsonify({"error": "Invalid OTP"}), 400

    if not company.get("otp_expiry") or datetime.now() > company["otp_expiry"]:
        return jsonify({"error": "OTP expired"}), 400

    # Clear OTP so it can't be reused
    cursor.execute("UPDATE companies SET otp=NULL, otp_expiry=NULL WHERE company_id=%s", (company["company_id"],))
    db.commit()

    return jsonify({"message": "✅ OTP verified", "company_id": company["company_id"], "email": company["email"]})


@app.route("/api/manager/resend-otp", methods=["POST"])
def manager_resend_otp():
    data = request.json or {}
    identifier = data.get("identifier")

    if not identifier:
        return jsonify({"error": "Identifier is required"}), 400

    cursor.execute("SELECT company_id, email FROM companies WHERE company_id=%s OR email=%s", (identifier, identifier))
    company = cursor.fetchone()

    if not company:
        return jsonify({"error": "Company not found"}), 404

    otp = random.randint(10000, 99999)
    otp_expiry = datetime.now() + timedelta(minutes=1)

    cursor.execute(
        "UPDATE companies SET otp=%s, otp_expiry=%s WHERE company_id=%s",
        (otp, otp_expiry, company["company_id"])
    )
    db.commit()

    ok = send_email(company["email"], "OTP for Password Reset (Resent)", f"Your new OTP is {otp}. Valid for 1 minute.")
    if not ok:
        return jsonify({"error": "Failed to send email"}), 500

    return jsonify({"message": "✅ Verification code resent", "company_id": company["company_id"], "email": company["email"]})

@app.route("/api/manager/reset-password", methods=["POST"])
def manager_reset_password():
    """
    Resets company (manager) password.
    Behaviour:
      - If companies.otp IS NULL in DB (OTP already verified / consumed), allow reset without providing otp.
      - If companies.otp IS NOT NULL, require otp parameter to match and be unexpired.
    """
    data = request.json or {}
    identifier = data.get("identifier")
    otp = data.get("otp")
    newPassword = data.get("newPassword")

    if not identifier or not newPassword:
        return jsonify({"error": "Identifier and newPassword are required"}), 400

    cursor.execute("SELECT * FROM companies WHERE company_id=%s OR email=%s", (identifier, identifier))
    company = cursor.fetchone()

    if not company:
        return jsonify({"error": "Company not found"}), 404

    db_otp = company.get("otp")
    db_expiry = company.get("otp_expiry")

    # If OTP present in DB -> require otp param and validate it
    if db_otp is not None:
        if otp is None:
            return jsonify({"error": "OTP is required"}), 400
        # compare as strings to avoid numeric formatting issues (leading zeros, etc.)
        if str(db_otp) != str(otp).strip():
            return jsonify({"error": "Invalid OTP"}), 400
        if not db_expiry or datetime.now() > db_expiry:
            return jsonify({"error": "OTP expired"}), 400

    # If db_otp is None => OTP was already verified (or not set) -> allow reset

    try:
        cursor.execute(
            "UPDATE companies SET password=%s, otp=NULL, otp_expiry=NULL WHERE company_id=%s",
            (newPassword, company["company_id"])
        )
        db.commit()
        return jsonify({"message": "✅ Password updated successfully!"})
    except Exception as e:
        print(f"ERROR updating password for company {company.get('company_id')}: {e}")
        return jsonify({"error": "Failed to update password"}), 500



# ==============================
# 🧠 Utility functions for OCR/CMM
# ==============================
def ocr_image(img, psm=4):
    """Performs OCR on a PIL image with a given Page Segmentation Mode."""
    gray = img.convert('L')
    gray = ImageEnhance.Contrast(gray).enhance(2)
    config = f'--oem 3 --psm {psm}'
    return pytesseract.image_to_string(gray, config=config).strip()

def _normalize_feature_no(s):
    """Keep only digits and hyphens (e.g., '140-1'); trim spaces."""
    if s is None:
        return ""
    return re.sub(r"[^0-9\-]", "", str(s)).strip()

def _pick_first_actual(v):
    s = "" if v is None else str(v).strip()
    if not s:
        return ""
    m = re.search(r'[+-]?\d+(?:\.\d+)?', s)
    if m:
        return m.group(0)
    return re.split(r'[,;|\s/]+', s)[0].strip()

def _to_float(v):
    if v is None:
        return None
    m = re.search(r'[+-]?\d+(?:\.\d+)?', str(v))
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None

def _is_numeric_like(v):
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).strip())
        return True
    except Exception:
        return False

def _is_feature_like(v):
    txt = _normalize_feature_no(v)
    return bool(txt) and (1 <= len(txt) <= 12)

def _is_refer_only(s: str) -> bool:
    """Return True if the feature text is marked as 'Refer Only'."""
    if s is None:
        return False
    t = str(s).replace("\u00A0", " ")
    t = re.sub(r"\s+", " ", t).strip().lower()
    t = re.sub(r"[.,;:()\[\]{}/_-]+", " ", t)
    return re.search(r"\bref(?:er(?:ence)?)?\s*only\b", t) is not None

# --- Supplier dataset bootstrapping
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
SUPPLIER_NAMES: List[str] = []

def _norm(s: str) -> str:
    return " ".join(str(s).split()).strip() if s is not None else ""

def _load_suppliers_on_boot() -> List[str]:
    names: List[str] = []

    # 1) JSON
    json_path = os.path.join(DATA_DIR, "suppliers.json")
    if os.path.isfile(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            if isinstance(payload, dict):
                if "suppliers" in payload and isinstance(payload["suppliers"], list):
                    names.extend([_norm(x) for x in payload["suppliers"]])
                elif "names" in payload and isinstance(payload["names"], list):
                    names.extend([_norm(x) for x in payload["names"]])
            elif isinstance(payload, list):
                names.extend([_norm(x) for x in payload])
        except Exception:
            pass

    # 2) Excel
    if not names:
        try:
            for fname in ("suppliers.xlsx", "suppliers.xlsm", "suppliers.xls"):
                x_path = os.path.join(DATA_DIR, fname)
                if os.path.isfile(x_path):
                    wb = openpyxl.load_workbook(x_path, read_only=True, keep_vba=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    header = [(_norm(h).lower() if h is not None else "") for h in rows[0]]
                    choices = ("supplier name", "supplier", "vendor", "vendor name", "name")
                    col_idx = 0
                    for i, h in enumerate(header):
                        if any(h.startswith(c) for c in choices):
                            col_idx = i
                            break
                    if col_idx == 0 and header and header[0] == "":
                        col_idx = 0
                    for r in rows[1:]:
                        if r and col_idx < len(r) and r[col_idx]:
                            names.append(_norm(r[col_idx]))
                    if names:
                        break
        except Exception:
            pass

    # 3) CSV
    if not names:
        c_path = os.path.join(DATA_DIR, "suppliers.csv")
        if os.path.isfile(c_path):
            try:
                with open(c_path, "r", encoding="utf-8", errors="ignore", newline="") as f:
                    rdr = csv.DictReader(f)
                    if rdr.fieldnames:
                        fn = [(_norm(h).lower() if h else "") for h in rdr.fieldnames]
                        choices = ("supplier name", "supplier", "vendor", "vendor name", "name")
                        col = None
                        for i, h in enumerate(fn):
                            if any(h.startswith(c) for c in choices):
                                col = rdr.fieldnames[i]
                                break
                        if col is None:
                            col = rdr.fieldnames[0]
                        for row in rdr:
                            val = row.get(col)
                            if val:
                                names.append(_norm(val))
            except Exception:
                pass

    if not names:
        names = [
            _norm("Acme Components"),
            _norm("Globex Engineering"),
            _norm("Initech Precision"),
            _norm("Umbrella Machining"),
        ]

    return sorted({n for n in names if n})

SUPPLIER_NAMES = _load_suppliers_on_boot()

def _find_header_row_and_cols(ws, max_scan_rows=50):
    feature_syns = [
        r"feature.?number",
        r"\bfeature\s*n(?:o|c)\b",
        r"\bfeature\s*nu",
        r"\bfeature\b",
        r"\bballoon\b",
        r"\bbubble\b",
        r"char(?:\.|acteristic)?\s*no",
        r"\bchar\s*no\b",
        r"\bdim(?:\.|ension)?\s*no\b",
        r"\bfeat\s*no\b",
        r"\bid\b"
    ]
    actual_syns = [
        r"\bactual\b",
        r"\bactua\b",
        r"\bmeasured\b",
        r"\bresult\b",
        r"\bvalue\b",
        r"\bmeasurement\b",
        r"\bread\b"
    ]

    best_row = None
    best_score = -1
    cand_feat_idx = None
    cand_act_idx = None

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > max_scan_rows:
            break
        row_vals = [str(c).strip().lower() if c is not None else "" for c in row]
        if not any(row_vals):
            continue

        feat_idx, act_idx = None, None
        score = 0
        for idx, cell in enumerate(row_vals):
            if not cell:
                continue
            if any(re.search(p, cell) for p in feature_syns):
                if feat_idx is None:
                    feat_idx = idx
                score += 2
            if any(re.search(p, cell) for p in actual_syns):
                if act_idx is None:
                    act_idx = idx
                score += 2
            if feat_idx is None and (cell.startswith("feature n") or cell.startswith("feature nu") or cell == "feature"):
                feat_idx = idx
                score += 1
            if act_idx is None and (cell.startswith("actua") or cell.startswith("meas") or cell.startswith("resul") or cell == "value"):
                act_idx = idx
                score += 1

        if feat_idx is not None and act_idx is not None and score > best_score:
            best_row, best_score = i, score
            cand_feat_idx, cand_act_idx = feat_idx, act_idx

    if cand_feat_idx is not None and cand_act_idx is not None:
        return best_row, cand_feat_idx, cand_act_idx

    col_stats = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > max_scan_rows:
            break
        vals = list(row)
        for idx, v in enumerate(vals):
            st = col_stats.setdefault(idx, {"feature_like": 0, "numeric_like": 0, "nonempty": 0})
            if v is not None and str(v).strip() != "":
                st["nonempty"] += 1
                if _is_feature_like(v):
                    st["feature_like"] += 1
                if _is_numeric_like(v):
                    st["numeric_like"] += 1

    feature_col = None
    actual_col = None
    best_feat = -1
    best_num = -1
    for idx, st in col_stats.items():
        if st["feature_like"] > best_feat and st["feature_like"] >= 3:
            best_feat = st["feature_like"]
            feature_col = idx
        if st["numeric_like"] > best_num and st["numeric_like"] >= 3:
            best_num = st["numeric_like"]
            actual_col = idx

    if feature_col is not None and actual_col is not None and feature_col != actual_col:
        return None, feature_col, actual_col

    return None, None, None


# ==============================
# 📄 OCR / EXTRACTION ENDPOINTS
# ==============================
@app.route('/api/extract-text', methods=['POST'])
def extract_text_from_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    filename = file.filename.lower()

    try:
        file_content = file.read()

        # --- Excel (IPS / generic) ---
        if filename.endswith(('.xlsm', '.xlsx', '.xls')):
            xlsm_bytes = io.BytesIO(file_content)
            wb = openpyxl.load_workbook(xlsm_bytes, data_only=True, keep_vba=True, read_only=True)

            # Adjust sheet name to your format if needed
            ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.active

            rows = list(ws.iter_rows(values_only=True))
            # Header expected at 2nd row based on your sample; adjust as needed
            if len(rows) < 2:
                return jsonify({"extracted_data": []})
            header = [h.strip().lower() if isinstance(h, str) else None for h in rows[1]]

            data = []
            for r in rows[2:]:
                rd = dict(zip(header, r))
                op = rd.get("operation")
                feat = rd.get("feature\nnumber")
                ref = rd.get("drawing ref")
                desc = rd.get("description")
                if op:
                    data.append({
                        "operation": str(op).strip(),
                        "feature_no": str(feat).strip() if feat else "",
                        "drawing_ref": str(ref).strip() if ref else "",
                        "description": str(desc).strip() if desc else "",
                    })
            return jsonify({"extracted_data": data})

        # --- PDF ---
        elif filename.endswith('.pdf'):
            pdf_doc = fitz.open(stream=file_content, filetype="pdf")

            # Case A: "inspection planning sheet" in filename
            if "inspection planning sheet" in filename:
                full_text = ""
                for page_num, page in enumerate(pdf_doc, start=1):
                    page_text = page.get_text("text").strip()
                    if len(page_text) < 50:
                        pix = page.get_pixmap(dpi=200)
                        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                        page_text = ocr_image(img, psm=4)
                    full_text += page_text + "\n"

                lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]
                # Group lines into rows starting with operation number
                grouped = []
                current = []
                for ln in lines:
                    if re.match(r"^\d{3,4}\b", ln):
                        if current:
                            grouped.append(" ".join(current))
                        current = [ln]
                    else:
                        current.append(ln)
                if current:
                    grouped.append(" ".join(current))

                final_data = []
                pattern = re.compile(r"^(\d{1,4})\s+(\S+)(?:\s+(SHEET[0-9A-Z\-]+))?\s+(.*)$")
                for row in grouped:
                    m = pattern.match(row)
                    if m:
                        op = m.group(1)
                        feature_no = m.group(2)
                        drawing_ref = m.group(3) or ""
                        desc = m.group(4).strip()
                        final_data.append({
                            "operation": op,
                            "feature_no": feature_no,
                            "drawing_ref": drawing_ref,
                            "description": desc
                        })
                if final_data:
                    return jsonify({"extracted_data": final_data})

            # Case B: fallback PDF text
            final_text = ""
            for page in pdf_doc:
                page_text = page.get_text().strip()
                if len(page_text) < 20:
                    pix = page.get_pixmap(dpi=200)
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    page_text = ocr_image(img, psm=4)
                final_text += page_text + "\n\n"
            return jsonify({"extracted_text": final_text})

        # --- DOCX ---
        elif filename.endswith('.docx'):
            doc = docx.Document(io.BytesIO(file_content))
            final_text = "\n".join([para.text for para in doc.paragraphs])
            return jsonify({"extracted_text": final_text})

        # --- Images ---
        elif filename.endswith(('.png', '.jpg', '.jpeg')):
            img = Image.open(io.BytesIO(file_content))
            final_text = ocr_image(img)
            return jsonify({"extracted_text": final_text})

        else:
            return jsonify({"error": "Unsupported file type"}), 400

    except Exception as e:
        print(f"ERROR in /api/extract-text: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Failed to process the file. Check server logs for details."}), 500


@app.route('/api/extract-text-stream', methods=['POST'])
def extract_text_stream():
    """
    Streams rows from a large Excel (xls/xlsx/xlsm) file as JSON array
    so the frontend can render rows incrementally.
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    filename = file.filename.lower()
    if not filename.endswith(('.xlsm', '.xlsx', '.xls')):
        return jsonify({"error": "Only Excel files are supported for streaming"}), 400

    def generate():
        xlsm_bytes = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(xlsm_bytes, data_only=True, keep_vba=True, read_only=True)
        ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.active

        # send JSON array start
        yield '{"extracted_data":['
        first = True

        rows = ws.iter_rows(values_only=True)
        try:
            next(rows)  # skip row 1
            header = [h.strip().lower() if isinstance(h, str) else None for h in next(rows)]  # header row 2
        except StopIteration:
            yield "]}"; return

        for r in rows:
            rd = dict(zip(header, r))
            op = rd.get("operation")
            feature_no_raw = rd.get("feature\nnumber")
            feature_no = str(feature_no_raw).strip() if feature_no_raw else ""
            # Only keep feature numbers that contain digits and no letters
            if not feature_no or not re.search(r'\d', feature_no) or re.search(r'[a-zA-Z]', feature_no):
                continue
            if not op:
                continue
            item = {
                "operation": str(op).strip(),
                "feature_no": feature_no,
                "drawing_ref": str(rd.get("drawing ref") or "").strip(),
                "description": str(rd.get("description") or "").strip()
            }
            if not first:
                yield ","
            first = False
            yield json.dumps(item)

        yield "]}"

               

    return Response(stream_with_context(generate()), mimetype='application/json')


@app.route('/api/ocr-image', methods=['POST'])
def ocr_cropped_image():
    if 'cropped_image' not in request.files:
        return jsonify({"error": "No cropped_image part in the request"}), 400

    file = request.files['cropped_image']
    if file.filename == '':
        return jsonify({"error": "No file selected for cropping"}), 400

    try:
        img = Image.open(file.stream)
        extracted_text = ocr_image(img, psm=6)
        return jsonify({"extracted_text": extracted_text})
    except Exception as e:
        print(f"ERROR in /api/ocr-image: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Failed to perform OCR on the cropped image."}), 500


# ==============================
# 📊 CMM Parser
# ==============================
@app.route('/api/parse-cmm', methods=['POST'])
def parse_cmm_report():
    """
    Accepts an Excel CMM measurement report and returns a merged mapping
    { normalized_feature_no: actual_value } across all sheets.
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    filename = file.filename.lower()
    if not filename.endswith(('.xlsm', '.xlsx', '.xls')):
        return jsonify({"error": "Only Excel files are supported"}), 400

    try:
        xls_bytes = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(xls_bytes, data_only=True, read_only=True, keep_vba=True)

        feature_map = {}        # key -> first/merged actual
        feature_range_map = {}  # key -> min/max float
        refer_only_keys = set()
        detector_debug = []     # to help debugging which columns were detected

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_idx, feat_col_idx, act_col_idx = _find_header_row_and_cols(ws)

            detector_debug.append({
                "sheet": sheet_name,
                "header_row_idx": header_row_idx,
                "feature_col_idx": feat_col_idx,
                "actual_col_idx": act_col_idx,
            })

            if feat_col_idx is None or act_col_idx is None:
                continue

            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if header_row_idx is not None and r_idx <= header_row_idx:
                    continue

                vals = list(row)
                feat_raw = vals[feat_col_idx] if feat_col_idx < len(vals) else None
                act_val = vals[act_col_idx] if act_col_idx < len(vals) else None
                key = _normalize_feature_no(feat_raw)

                if _is_refer_only(feat_raw) and key:
                    refer_only_keys.add(key)
                if not key:
                    continue

                val = _pick_first_actual(act_val)
                if key not in feature_map or (not feature_map[key] and val):
                    feature_map[key] = val

                fv = _to_float(val)
                if fv is not None:
                    rng = feature_range_map.get(key)
                    if rng is None:
                        feature_range_map[key] = {"min": fv, "max": fv}
                    else:
                        if fv < rng["min"]:
                            rng["min"] = fv
                        if fv > rng["max"]:
                            rng["max"] = fv

        if not feature_map:
            return jsonify({
                "error": "Could not detect 'Feature Number' and 'Actual' columns in any sheet",
                "detector_debug": detector_debug
            }), 400

        return jsonify({
            "feature_actual_map": feature_map,
            "feature_range_map": feature_range_map,
            "refer_only_keys": sorted(refer_only_keys),
            "detector_debug": detector_debug
        })

    except Exception as e:
        print(f"ERROR in /api/parse-cmm: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Failed to parse CMM report."}), 500


# ==============================
# 🏷️ Supplier Autocomplete
# ==============================
@app.route("/api/suppliers", methods=["GET"])
def suppliers():
    try:
        prefix = request.args.get("prefix", "", type=str) or ""
        limit = request.args.get("limit", default=10, type=int)
        p = _norm(prefix).lower()
        matches = [s for s in SUPPLIER_NAMES if s.lower().startswith(p)] if p else SUPPLIER_NAMES[:]
        if limit is not None and limit >= 0:
            matches = matches[:limit]
        return jsonify({"suppliers": matches})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================
# 📥 Save Form1 PDF + Excel for Employee
# ==============================
@app.route("/api/employees/save-form1", methods=["POST"])
def save_form1_files():
    try:
        employee_id = request.form.get("employee_id")
        pdf_file = request.files.get("pdf")
        excel_file = request.files.get("excel")

        if not employee_id or not pdf_file or not excel_file:
            return jsonify({"error": "❌ employee_id, pdf, and excel are required"}), 400

        pdf_data = pdf_file.read()
        excel_data = excel_file.read()

        cursor.execute(
            "UPDATE employees SET form1_pdf=%s, form1_excel=%s WHERE employee_id=%s",
            (pdf_data, excel_data, employee_id)
        )
        db.commit()

        return jsonify({"message": f"✅ Form1 PDF & Excel stored for employee {employee_id}"})
    except Exception as e:
        print("❌ Error saving Form1 files:", e)
        return jsonify({"error": str(e)}), 500


# ==============================
# 🚀 START SERVER
# ==============================
if __name__ == "__main__":
    app.run(port=5000, debug=True)
